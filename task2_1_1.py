import csv
import math
import re
import os
from operator import itemgetter
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Border, NamedStyle, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


input_sentences = {'name': 'Введите название файла: ', 'job_name': 'Введите название профессии: '}
currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class DataSet:
    def __init__(self, file_name: str, vacancies_objects: list):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def read_file(self):
        file = open(self.file_name, encoding='utf_8_sig')
        reader = csv.reader(file)
        vacancies = [line for line in reader]
        column_names = vacancies.pop(0)
        return column_names, [job for job in vacancies if len(job) == len(column_names) and job.count('') == 0]

    def get_reformed_file(self, reader: list, list_naming: list):
        descriptions = []
        for vac in reader:
            description = dict()
            for i in range(len(list_naming)):
                vac[i] = re.sub(r'\<[^>]*\>', '', vac[i])
                if vac[i].count('\n') != 0:
                    vac[i] = ", ".join(vac[i].split("\n"))
                vac[i] = ' '.join(vac[i].split())
                description[list_naming[i]] = vac[i]
            descriptions.append(Vacancy(description))
        return descriptions


class Vacancy:
    def __init__(self, descriptions: dict):
        self.name = descriptions['name']
        self.salary = Salary(descriptions)
        self.area_name = descriptions['area_name']
        self.published_at = int(descriptions['published_at'][:4])


class Salary:
    def __init__(self, descriptions: dict):
        self.salary_from = self.convert_string_to_int(descriptions['salary_from'])
        self.salary_to = self.convert_string_to_int(descriptions['salary_to'])
        self.salary_currency = descriptions['salary_currency']
        self.salary_in_rub = self.convert_to_rubles((self.salary_to + self.salary_from) / 2, self.salary_currency)

    def convert_to_rubles(self, average_salary, currency):
        return average_salary * currency_to_rub[currency]

    def convert_string_to_int(self, line: str):
        return int(line.split('.')[0])


class InputConnect:
    def __init__(self, sentences: dict):
        self.name = input(sentences['name'])
        self.job_name = input(sentences['job_name'])
        self.vacancies_info = {}
        self.vacancies_info_names = ['Динамика уровня зарплат по годам',
                                     'Динамика количества вакансий по годам',
                                     'Динамика уровня зарплат по годам для выбранной профессии',
                                     'Динамика количества вакансий по годам для выбранной профессии',
                                     'Уровень зарплат по городам (в порядке убывания)',
                                     'Доля вакансий по городам (в порядке убывания)']

    def fill_vacancies_info(self, vacancies: list):
        years = set()
        cities = dict()
        for vac in vacancies:
            years.add(vac.published_at)
            if vac.area_name in cities:
                cities[vac.area_name].append(vac.salary.salary_in_rub)
            else:
                cities[vac.area_name] = [vac.salary.salary_in_rub]
        years = sorted(years)
        years = list(range(min(years), max(years) + 1))
        cities = [city for city in cities.items() if math.floor(len(city[1]) / len(vacancies) * 100) >= 1]
        vacancies_by_year_and_city = self.get_vacancies_info_by_year(vacancies, years) + \
                                     self.get_vacancies_info_by_city(vacancies, cities)
        for i in range(len(vacancies_by_year_and_city)):
            self.vacancies_info[self.vacancies_info_names[i]] = vacancies_by_year_and_city[i]

    def get_vacancies_info_by_year(self, vacancies: list, years: list):
        all_salaries_by_year = {year: [] for year in years}
        all_vacancies_count_by_year = {year: 0 for year in years}
        exact_salaries_by_year = {year: [] for year in years}
        exact_vacancies_count_by_year = {year: 0 for year in years}
        for vac in vacancies:
            all_salaries_by_year[vac.published_at].append(vac.salary.salary_in_rub)
            all_vacancies_count_by_year[vac.published_at] += 1
            if self.job_name in vac.name:
                exact_salaries_by_year[vac.published_at].append(vac.salary.salary_in_rub)
                exact_vacancies_count_by_year[vac.published_at] += 1
        all_salaries_by_year = {year: int(sum(salaries) / len(salaries)) if len(salaries) != 0 else 0
                                for year, salaries in all_salaries_by_year.items()}
        exact_salaries_by_year = {year: int(sum(salaries) / len(salaries)) if len(salaries) != 0 else 0
                                  for year, salaries in exact_salaries_by_year.items()}
        return all_salaries_by_year, all_vacancies_count_by_year, exact_salaries_by_year, exact_vacancies_count_by_year

    def get_vacancies_info_by_city(self, vacancies: list, cities):
        all_fractions_by_city = {city[0]: round(len(city[1]) / len(vacancies), 4) for city in cities}
        all_fractions_by_city = dict(sorted(all_fractions_by_city.items(),
                                            key=itemgetter(1), reverse=True)[:10])
        cities = sorted(cities, key=lambda salaries: sum(salaries[1]) / len(salaries[1]), reverse=True)
        all_salaries_by_cities = {city[0]: int(sum(city[1]) / len(city[1])) if len(city[1]) > 0 else 0
                                  for city in cities[:min(10, len(cities))]}
        return all_salaries_by_cities, all_fractions_by_city

    def print_vacancies_info(self, vacancies: list, table_name: str, years_stats_name: str, area_stats_name: str):
        self.fill_vacancies_info(vacancies)
        for key, value in self.vacancies_info.items():
            print(f"{key}: {value}")
        rep = Report(table_name, self.vacancies_info, self.job_name)
        rep.generate_excel(years_stats_name, area_stats_name)


class Report:
    def __init__(self, name, years_data, job_name):
        self.name = name
        self.years_data = years_data
        self.job_name = job_name

    def reform_cells(self, current_sheet, table_style: NamedStyle, sheet_data: list, highlight: Font):
        dimensions = {}
        for i in range(1, len(sheet_data) + 2):
            for cell in current_sheet[i]:
                if cell.value == '':
                    dimensions[cell.column_letter] = 0
                else:
                    cell.style = table_style
                    if i == 1:
                        cell.font = highlight
                    dimensions[cell.column_letter] = max(dimensions.get(cell.column_letter, 0), len(str(cell.value)))
        for key, value in dimensions.items():
            current_sheet.column_dimensions[key].width = value + 2

    def create_style(self):
        new_style = NamedStyle('highlight')
        new_style.font = Font(name='Calibri', size=11, color='000000')
        border = Side(style="thin", color="000000")
        new_style.border = Border(left=border, right=border, top=border, bottom=border)
        return new_style

    def generate_excel(self, years_stats_name: str, area_stats_name: str):
        table = Workbook()
        table_style = self.create_style()
        highlight = Font(bold=True)
        self.create_years_statistics(table, table_style, highlight, years_stats_name)
        self.create_area_statistics(table, table_style, highlight, area_stats_name)
        table.remove(table.active)
        table.save(self.name)

    def create_years_statistics(self, table: Workbook, table_style: NamedStyle, highlight: Font, sheet_name: str):
        year_stats = table.create_sheet(sheet_name)
        years = list(self.years_data['Динамика уровня зарплат по годам'].keys())
        year_stats.append(['Год', 'Средняя зарплата', f"Средняя зарплата - {self.job_name}", 'Количество вакансий',
                           f"Количество вакансий - {self.job_name}"])
        for year in years:
            year_stats.append([year, self.years_data['Динамика уровня зарплат по годам'][year],
                               self.years_data['Динамика уровня зарплат по годам для выбранной профессии'][year],
                               self.years_data['Динамика количества вакансий по годам'][year],
                               self.years_data['Динамика количества вакансий по годам для выбранной профессии'][year]])
        self.reform_cells(year_stats, table_style, years, highlight)

    def create_area_statistics(self, table: Workbook, table_style: NamedStyle, highlight: Font, sheet_name: str):
        area_stats = table.create_sheet(sheet_name)
        city_salaries = list(self.years_data['Уровень зарплат по городам (в порядке убывания)'].keys())
        cities_fractions = list(self.years_data['Доля вакансий по городам (в порядке убывания)'].keys())
        area_stats.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for i in range(len(city_salaries)):
            area_stats.append([city_salaries[i],
                               self.years_data['Уровень зарплат по городам (в порядке убывания)'][city_salaries[i]],
                               '', cities_fractions[i],
                               self.years_data['Доля вакансий по городам (в порядке убывания)'][cities_fractions[i]]])
        self.reform_cells(area_stats, table_style, city_salaries, highlight)
        for cell in area_stats['E']:
            cell.number_format = FORMAT_PERCENTAGE_00


csv_file = InputConnect(input_sentences)
if os.path.getsize(csv_file.name) == 0:
    print('Пустой файл')
else:
    data = DataSet(csv_file.name, [])
    info = data.read_file()
    data.vacancies_objects = data.get_reformed_file(info[1], info[0])
    csv_file.print_vacancies_info(data.vacancies_objects, 'report.xlsx','Статистика по годам', 'Статистика по городам')
